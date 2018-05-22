#-*- coding: UTF-8 -*- 
#!/usr/bin/python
from openpyxl import load_workbook
import re
import os
from log import logger
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *


class ObjectDict(dict):
    """Makes a dictionary behave like an object, with attribute-style access.
    """
    def __getattr__(self, name):
        try:
            return self[name]
        except KeyError:
            raise AttributeError(name)

    def __setattr__(self, name, value):
        self[name] = value


class AreaCodeItem(ObjectDict):

    def __init__(self, code, province, city, district):
        super(AreaCodeItem, self).__init__()
        self.code = code
        self.province = province
        self.city = city
        self.district = district

    def __unicode__(self):
        return u"{0}, {1}, {2}, {3}".format(self.code, self.province, self.city, self.district)

    def __str__(self):
        return unicode(self).encode("utf-8")


class AreaCode(object):

    def __init__(self):
        super(AreaCode, self).__init__()
        self.wb = load_workbook("data/code.xlsx")
        self.sheet1 = self.wb[self.wb.sheetnames[0]]

        self.districtSuffixs = [u"区", u"县", u"市", u"旗", u"镇"]
        self.citySuffixs = [u"市", u"县", u'自治州', u'自治县']
        self.city_defaultSuffix = u"市"
        self.provinceSuffix = u"省"


        self._items = []
        self._itemsCodes = {}

        self.provinceKeys = set()
        self.cityKeys = set()
        self.districtKeys = set()


        self.city_districts = {}
        self.province_districts = {}

        rows = self.sheet1.rows
        i = 0
        for row in rows:
            # print i, row, (i==0)
            if i > 0:
                row_values = []
                for cell in row:
                    row_values.append(cell.value)
                item = AreaCodeItem(row_values[0], row_values[1], row_values[2], row_values[3])
                self._items.append(item)

                if item.district:
                    self._itemsCodes[item.district] = item.code

                    self.province_districts[item.province].add(item.district)
                    self.province_districts[item.province + self.provinceSuffix].add(item.district)

                    self.districtKeys.add(item.district)

                    for suffix in self.districtSuffixs:
                        if item.district.endswith(suffix):
                            district = item.district[0: len(item.district) - 1]
                            self.districtKeys.add(district)
                            self._itemsCodes[district] = item.code

                    # self.province_districts[item.province].add(district)
                    # self.province_districts[item.province + self.provinceSuffix].add(district)


                    cityName = item.city
                    for suffix in self.citySuffixs:
                        if suffix in item.city:
                            continue
                        else:
                            cityName = item.city + self.city_defaultSuffix
                    self.city_districts[item.city].add(item.district)
                    self.city_districts[cityName].add(item.district)

                elif item.city and item.district is None:
                    cityName = item.city

                    for suffix in self.citySuffixs:
                        if suffix in item.city:
                            continue
                        else:
                            cityName = item.city + self.city_defaultSuffix

                    self._itemsCodes[cityName] = item.code
                    self._itemsCodes[item.city] = item.code

                    self.cityKeys.add(cityName)
                    self.cityKeys.add(item.city)


                    self.city_districts[cityName] = set()
                    self.city_districts[item.city] = set()


                elif item.province and item.city is None and item.district is None:
                    self._itemsCodes[item.province + self.provinceSuffix] = item.code
                    self._itemsCodes[item.province] = item.code


                    self.provinceKeys.add(item.province)
                    self.provinceKeys.add(item.province + self.provinceSuffix)


                    self.province_districts[item.province] = set()
                    self.province_districts[item.province + self.provinceSuffix] = set()


            i = i + 1


    def getCodeByCodeArea(self, district):
        code = 0
        if (isinstance(district, unicode) and self._itemsCodes.has_key(district)):
            code = self._itemsCodes.get(district, 0)
        elif (isinstance(district, str) and self._itemsCodes.has_key(district.decode("utf-8"))):
            code = self._itemsCodes.get(district.decode("utf-8"), 0)
        return code


    def getCodeAreaByAddress(self, address, defautProvice=""):

        # u"(?<=市).*?(?=区)"
        # logger.info("1")
        for citySuffix in self.citySuffixs:
            for districtSuffix in self.districtSuffixs:
                reg = u"(?<={}).*?(?={})".format(citySuffix, districtSuffix)
                matchObj = re.search(reg, address)
                if matchObj:
                    ret = matchObj.group() + districtSuffix
                    if ret in self.districtKeys:
                        return ret
        # logger.info("2")
        # u"(?<=省).*?(?=区)"
        for districtSuffix in self.districtSuffixs:
                reg = u"(?<={}).*?(?={})".format(self.provinceSuffix, districtSuffix)
                matchObj = re.search(reg, address)
                if matchObj:
                    ret = matchObj.group() + districtSuffix
                    if ret in self.districtKeys:
                        return ret
        # logger.info("3")
        #余杭区/杭州余杭区
        for districtSuffix in self.districtSuffixs:
            reg = u".*?(?={})".format(districtSuffix)
            matchObj = re.search(reg, address)
            if matchObj:
                ret = matchObj.group() + districtSuffix
                if ret in self.districtKeys:
                    return ret
                else:
                    results = []
                    for d in self.districtKeys:
                        if ret.endswith(d):
                            return d
                        if d in matchObj.group():
                            results.append(d)

                    l = 0
                    ret = ""
                    for item in results:
                        if len(item) >= l:
                            l = len(item)
                            ret = item
                    if ret and (ret + districtSuffix) in self.districtKeys:

                        for city in self.cityKeys:
                            if city in address:
                                if  (ret + districtSuffix) in self.city_districts[city]:
                                    return ret + districtSuffix

        # logger.info("4")
        #杭州余杭xxxxxx/杭州市余杭xxxxxx
        results = []
        hasCity = False

        if defautProvice:
            districtKeys = self.province_districts[defautProvice]
        else:
            districtKeys = self.districtKeys

        for districtKey in districtKeys:
            # print districtKey, address
            districtName = districtKey[0:len(districtKey) - 1]
            if districtName in address:
                # print districtName
                for districtSuffix in self.districtSuffixs:
                    if (districtName+districtSuffix) in districtKeys:
                        for cityKey in self.cityKeys:
                            if cityKey in address:
                                hasCity = True
                                if (districtName+districtSuffix) in self.city_districts[cityKey]:
                                    return districtName+districtSuffix

                        if not hasCity:
                            results.append(districtName+districtSuffix)
        if results:
            l = 0
            ret = ""
            for item in results:
                if len(item) >= l:
                    l = len(item)
                    ret = item
            if ret and ret in districtKeys:
                return ret

        # logger.info("5")
        #杭州市xxxxxx

        for citySuffix in self.citySuffixs:
            reg = u".*?(?={})".format(citySuffix)
            matchObj = re.search(reg, address)
            if matchObj:
                ret = matchObj.group() + citySuffix
                if ret in self.cityKeys:
                    return ret

        # logger.info("6")
        #杭州xxxx
        for cityKey in self.cityKeys:
            if cityKey in address:
                for citySuffix in self.citySuffixs:
                    if (cityKey+citySuffix) in self.cityKeys:
                        return cityKey+citySuffix
        # logger.info("7")
        #浙江省
        matchObj = re.search(u".*?(?=省)", address)
        if matchObj:
            ret = matchObj.group() + self.provinceSuffix
            if ret in self.provinceKeys:
                return ret

        # logger.info("8")
        #浙江

        for provinceKey in self.provinceKeys:
            if provinceKey in address:
                if (provinceKey + self.provinceSuffix) in self.provinceKeys:
                    return (provinceKey + self.provinceSuffix)

        return defautProvice

    def getProvices(self):
        provinces = set()
        for item in self._items:
            provinces.add(item.province)
        return provinces

    def testCodeArea(self):
        print self.getCodeAreaByAddress(u"萧山区萧山经济技术开发区启迪路198号C-310-1室")
        print self.getCodeAreaByAddress(u"杭州市西湖区西园八路11号1幢1楼113室")
        print self.getCodeAreaByAddress(u"桐庐县城白云源路877号新天地花园21幢2201室")
        print self.getCodeAreaByAddress(u"杭州市转塘街道美院南街99号3476室")
        print self.getCodeAreaByAddress(u"浙江省桐庐县城南街道仁智村天井坞68号")
        print self.getCodeAreaByAddress(u"浙江省城南街道仁智村天井坞68号", u"浙江省")
        print self.getCodeAreaByAddress(u"城南街道仁智村天井坞68号", u"浙江省")
        print self.getCodeAreaByAddress(u"杭州市余杭区良渚街道时代大厦828室")
        print self.getCodeAreaByAddress(u"杭州余杭区良渚街道行宫塘村二组")
        print self.getCodeAreaByAddress(u"杭州市文三路199号创业大厦0657—1室")
        print self.getCodeAreaByAddress(u"萧山经济技术开发区天辰国际广场4幢811室", u"浙江省")
        print self.getCodeAreaByAddress(u"萧山高新科技广场1幢1楼", u"浙江省")

    def testCode(self):
        print self.getCodeByCodeArea(u"杭州市")




class ExcelHandler(QObject):

    areasFinished = pyqtSignal(list)
    codesFinished = pyqtSignal(list)
    excelstarted = pyqtSignal(unicode)
    excelsFininshed = pyqtSignal(list)

    execelError = pyqtSignal(unicode)

    def __init__(self):
        super(ExcelHandler, self).__init__()
        self.areaCode = AreaCode()
        # self.areaCode.testCode()
        # self.areaCode.testCodeArea()


    def handleExecl(self, excelFile, outputDir="", defautProvice=u""):
        logger.info("%s, %s", excelFile, os.path.exists(excelFile))
        outputFile = ""
        if os.path.exists(excelFile):
            self.excelstarted.emit(excelFile)
            filename = os.path.basename(excelFile)
            self.wb = load_workbook(excelFile)
            self.sheet1 = self.wb[self.wb.sheetnames[0]]
            for i in range(2,self.sheet1.max_row + 1):
                address = self.sheet1.cell(row=i, column=5).value
                if address and len(address) > 0:
                    codeArea = self.areaCode.getCodeAreaByAddress(address, defautProvice=self.wb.sheetnames[0])
                    code = self.areaCode.getCodeByCodeArea(codeArea)

                    self.sheet1["I{}".format(i)] = code

                    phone = self.sheet1.cell(row=i, column=4).value
                    phones = []
                    for item in phone.split(";"):
                        if len(item.strip()) == 11:
                            phones.append(item.strip())
                    self.sheet1["D{}".format(i)] = ";".join(phones)
                else:
                    self.execelError.emit(excelFile)
                    return

            if len(outputDir) == 0:
                outputDir = "{}{}{}".format( os.path.dirname(os.path.abspath(__file__)), os.sep, "output")
                if not os.path.exists(outputDir):
                    os.mkdir(outputDir)
            outputFile = u"{}{}{}".format(outputDir, os.sep, filename)
            self.wb.save(outputFile)
        return outputFile

    def handleExecls(self, excelFiles, outputDir="", defautProvice=u""):
        outputFiles = []
        for excelFile in excelFiles:
            f = self.handleExecl(excelFile, outputDir, defautProvice)
            if f:
                outputFiles.append(f)
        if outputFiles:
            self.excelsFininshed.emit(outputFiles)

    def handleAddress(self, addresss, defautProvice=u""):
        results = []
        for address in addresss:
            code = self.areaCode.getCodeAreaByAddress(address, defautProvice)
            results.append(code)
            # logger.info("%s %s", code, address)
        self.areasFinished.emit(results)
        return results

    def getCodesByAreas(self, areas):
        results = [str(self.areaCode.getCodeByCodeArea(area)) for area in areas]
        self.codesFinished.emit(results)
        return results


class DropArea(QLabel):

    changed = pyqtSignal(list)

    def __init__(self, parent = None):
        super(DropArea, self).__init__(parent)

        self.setMinimumSize(200, 200)
        self.setFrameStyle(QFrame.Sunken | QFrame.StyledPanel)
        self.setAlignment(Qt.AlignCenter)
        self.setAcceptDrops(True)
        self.setAutoFillBackground(True)
        self.clear()
        self.setStyleSheet("color: green; font-size: 16px;")

    def dragEnterEvent(self, event):
        self.setText("拖动Excel文件到此一键转换")
        self.setBackgroundRole(QPalette.Highlight)
        event.acceptProposedAction()


    def dragMoveEvent(self, event):
        event.acceptProposedAction()

    def dropEvent(self, event):
        mimeData = event.mimeData()
        if mimeData.hasUrls():
            self.setText(u"开始转换中(支持格式：.xlsx,.xlsm,.xltx,.xltm)")
            self.changed.emit([url.path() for url in mimeData.urls()])
        else:
            self.setText("Cannot display data")

        self.setBackgroundRole(QPalette.Dark)
        event.acceptProposedAction()

    def dragLeaveEvent(self, event):
        self.clear()
        event.accept()

    def mouseDoubleClickEvent(self, event):
        output = "{}{}{}".format( os.path.dirname(os.path.abspath(__file__)), os.sep, "output")
        QDesktopServices().openUrl(QUrl(output))
        event.accept()

    def clear(self):
        self.setText("拖动Excel文件到此一键转换")
        self.setBackgroundRole(QPalette.Dark)


class MainWindow(QLabel):

    addresssChanged = pyqtSignal(list, unicode)
    areasChanged = pyqtSignal(list, unicode)
    excelAdded = pyqtSignal(list, unicode)

    def __init__(self):
        super(MainWindow, self).__init__()
        self.initUI()
        self.initController()
        self.initConnect()


    def initUI(self):
        self.setWindowIcon(QIcon("icons/app.ico"))
        self.setWindowTitle(u"区码转换工具")
        self.setMinimumSize(1000, 600)

        self.provinceLable = QLabel(u"输入省份")
        self.proviceLineText = QLineEdit()
        self.proviceLineText.setPlaceholderText(u"输入正确的省份名字可以提高准确率（例如：浙江、浙江省）")
        self.clearButton = QPushButton(u"清空")

        self.addressLabel = QLabel(u"输入地址")

        self.addressEdit = QTextEdit()
        self.addressEdit.setAcceptRichText(False)

        self.areaLabel = QLabel(u"标准区名")
        self.areaEdit = QTextEdit()
        self.areaEdit.setAcceptRichText(False)

        self.codeLabel = QLabel(u"区码")
        self.codeEdit = QTextEdit()
        self.codeEdit.setAcceptRichText(False)

        self.dropArea = DropArea()

        self.statusLable = QLabel()

        proviceLayout = QHBoxLayout()
        proviceLayout.addWidget(self.provinceLable)
        proviceLayout.addWidget(self.proviceLineText, 1)
        proviceLayout.addWidget(self.clearButton)

        gridLayout = QGridLayout()
        gridLayout.addWidget(self.addressLabel, 0, 0, Qt.AlignLeft)
        gridLayout.addWidget(self.addressEdit, 1, 0, Qt.AlignLeft)
        gridLayout.addWidget(self.areaLabel, 0, 1, Qt.AlignLeft)
        gridLayout.addWidget(self.areaEdit, 1, 1, Qt.AlignLeft)
        gridLayout.addWidget(self.codeLabel, 0, 2, Qt.AlignLeft)
        gridLayout.addWidget(self.codeEdit, 1, 2, Qt.AlignLeft)
        # gridLayout.setColumnStretch(0, 100)
        gridLayout.setSpacing(0)
        gridLayout.setContentsMargins(0, 0, 0, 0)


        mainLayout = QVBoxLayout()
        mainLayout.addLayout(proviceLayout)
        mainLayout.addLayout(gridLayout)
        mainLayout.addWidget(self.dropArea)
        mainLayout.addWidget(self.statusLable)
        mainLayout.setContentsMargins(10, 10, 10, 10)
        self.setLayout(mainLayout)

        self.setStyleSheet("QLabel{ color: green; font-size: 16px;}")

    def initController(self):
        self.excelHandler = ExcelHandler()

        self.execlThread = QThread()
        self.excelHandler.moveToThread(self.execlThread)
        self.execlThread.start()

        self.areasTimer = QTimer()
        self.areasTimer.setSingleShot(True)
        self.areasTimer.setInterval(500)

        self.codeTimer = QTimer()
        self.codeTimer.setSingleShot(True)
        self.codeTimer.setInterval(500)

    def initConnect(self):
        self.clearButton.clicked.connect(self.clearContent)

        self.addressEdit.textChanged.connect(self.areasTimer.start)
        self.areaEdit.textChanged.connect(self.codeTimer.start)

        self.dropArea.changed.connect(self.handleDropChanged)

        self.areasTimer.timeout.connect(self.onAreasTimerTimeOut)
        self.codeTimer.timeout.connect(self.onCodeTimerTimeOut)

        self.addresssChanged.connect(self.excelHandler.handleAddress)
        self.areasChanged.connect(self.excelHandler.getCodesByAreas)
        self.excelAdded.connect(self.excelHandler.handleExecls)

        self.excelHandler.areasFinished.connect(self.handleAreasFinished)
        self.excelHandler.codesFinished.connect(self.handleCodesFinished)
        self.excelHandler.excelstarted.connect(self.onExcelstarted)
        self.excelHandler.excelsFininshed.connect(self.execlsFinishd)
        self.excelHandler.execelError.connect(self.showExecelError)

    def resizeEvent(self, event):
        self.areaEdit.setFixedWidth((event.size().width() / 4))
        self.codeEdit.setFixedWidth((event.size().width() / 4))
        self.addressEdit.setFixedWidth((event.size().width() / 2) - 20)

    def clearContent(self):
        self.addressEdit.clear()
        self.areaEdit.clear()
        self.codeEdit.clear()
        self.statusLable.setText(u"")

    def onAreasTimerTimeOut(self):
        self.areaEdit.clear()
        self.codeEdit.clear()
        data = self.addressEdit.toPlainText().strip()
        if len(data) > 0:
            self.statusLable.setText(u"正在转换中......")
            addresss = data.split("\n")
            self.addresssChanged.emit(addresss, self.proviceLineText.text().strip())
        else:
            self.statusLable.setText(u"")

    def onCodeTimerTimeOut(self):
        data = self.areaEdit.toPlainText()
        if len(data) > 0:
            areas = data.split("\n")
            self.areasChanged.emit(areas, self.proviceLineText.text().strip())

    def handleAreasFinished(self, areas):
        self.areaEdit.setText(u"\n".join(areas))
        self.statusLable.setText(u"转换完成, 一共{}行数据".format(len(areas)))

    def handleCodesFinished(self, codes):
        self.codeEdit.setText(u"\n".join(codes))

    def handleDropChanged(self, urls):
        vaildUrls = []
        for url in urls:
            if os.path.splitext(url)[1].endswith("xlsx"):
                vaildUrls.append(url)
        self.excelAdded.emit(vaildUrls, self.proviceLineText.text().strip())

    def onExcelstarted(self, excel):
        self.dropArea.setText(u"正在转换：{}".format(os.path.basename(excel)))

    def execlsFinishd(self, urls):
        self.dropArea.setText(u"转换成功，双击打开查看转换文件")

    def showExecelError(self, excel):
        self.dropArea.setText(u"{}格式有误".format(os.path.basename(excel)))


if __name__ == '__main__':
    import sys
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
